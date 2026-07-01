from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QScrollArea,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QFileDialog, QMessageBox,
)
from PyQt6.QtGui import QColor, QKeySequence, QShortcut
from PyQt6.QtCore import Qt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from src.moyennes import COLS_MOYENNE, COLS_CV, BLOCS_RESUME, RMR_LABELS, deduire_visite, calculer_rmr

COULEURS_BLOCS = {
    'last4':       QColor(221, 235, 247),
    'stable_vo2':  QColor(226, 239, 218),
    'stable_peto2': QColor(252, 228, 214),
    'vo2_min':     QColor(226, 214, 236),
}

COULEURS_EXCEL = {
    'last4':        'FFDDEBF7',
    'stable_vo2':   'FFE2EFDA',
    'stable_peto2': 'FFFCE4D6',
    'vo2_min':      'FFE2D6EC',
}


class PageResume(QWidget):
    def __init__(self, resume_data: dict):
        super().__init__()
        self.resume_data = resume_data
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        header = QWidget()
        header.setStyleSheet("background-color: #f5f5f5;")
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(32, 28, 32, 16)

        row_titre = QHBoxLayout()
        title = QLabel("Résumé")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #2b2d3b;")
        row_titre.addWidget(title)
        row_titre.addStretch()

        self.btn_supprimer = QPushButton("Supprimer la sélection")
        self.btn_supprimer.setStyleSheet(self._style_btn_danger())
        self.btn_supprimer.clicked.connect(self._supprimer_selection)
        row_titre.addWidget(self.btn_supprimer)

        self.btn_vider = QPushButton("Vider le résumé")
        self.btn_vider.setStyleSheet(self._style_btn_danger())
        self.btn_vider.clicked.connect(self._vider_resume)
        row_titre.addWidget(self.btn_vider)

        self.btn_export = QPushButton("Exporter en Excel")
        self.btn_export.setStyleSheet(self._style_btn_vert())
        self.btn_export.clicked.connect(self._exporter_excel)
        row_titre.addWidget(self.btn_export)

        header_layout.addLayout(row_titre)
        outer.addWidget(header)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        container = QWidget()
        scroll.setWidget(container)
        layout = QVBoxLayout(container)
        layout.setContentsMargins(32, 16, 32, 32)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

        shortcut_del = QShortcut(QKeySequence.StandardKey.Delete, self.table)
        shortcut_del.activated.connect(self._supprimer_selection)
        self.table.setStyleSheet("""
            QTableWidget { color: #2b2d3b; background-color: white;
                           border: 1px solid #ddd; border-radius: 4px; }
            QTableWidget::item:selected { background-color: #c8d8f8; color: #2b2d3b; }
            QHeaderView::section {
                background-color: #e8eaf0; color: #2b2d3b;
                padding: 4px 8px; border: none;
                border-right: 1px solid #ccc; font-weight: bold; font-size: 11px;
            }
        """)
        layout.addWidget(self.table)
        outer.addWidget(scroll)

    def showEvent(self, event):
        super().showEvent(event)
        self._actualiser_tableau()

    # ── Construction du tableau ────────────────────────────────────────────────

    def _vars_bloc(self) -> list:
        """Liste (clé_stats, libellé, décimales) communes à chaque bloc."""
        vars_ = [(c, c, 3 if c == 'Q.R.' else 2) for c in COLS_MOYENNE]
        vars_ += [(cle, lbl, 3) for cle, lbl in COLS_CV]
        return vars_

    def _colonnes(self) -> list:
        cols = ["Examen", "Visite"]
        for _, prefixe in BLOCS_RESUME:
            for _, lbl, _ in self._vars_bloc():
                cols.append(f"{prefixe}-{lbl}")
        for cle, _ in BLOCS_RESUME:
            cols.append(RMR_LABELS[cle])
        return cols

    def _actualiser_tableau(self):
        noms = list(self.resume_data.keys())
        vars_bloc = self._vars_bloc()
        col_headers = self._colonnes()

        self.table.clear()
        self.table.setColumnCount(len(col_headers))
        self.table.setRowCount(len(noms))
        self.table.setHorizontalHeaderLabels(col_headers)

        for row, nom in enumerate(noms):
            res = self.resume_data[nom]
            col = 0

            self.table.setItem(row, col, QTableWidgetItem(nom))
            col += 1

            visite = deduire_visite(nom)
            self.table.setItem(row, col, QTableWidgetItem(str(visite) if visite else ""))
            col += 1

            for cle_bloc, _ in BLOCS_RESUME:
                periode = res.get(cle_bloc)
                stats = periode['stats'] if periode else {}
                couleur = COULEURS_BLOCS[cle_bloc]
                for cle_stat, _, decimales in vars_bloc:
                    val = stats.get(cle_stat)
                    item = QTableWidgetItem(f"{val:.{decimales}f}" if val is not None else "—")
                    item.setBackground(couleur)
                    self.table.setItem(row, col, item)
                    col += 1

            for cle_bloc, _ in BLOCS_RESUME:
                periode = res.get(cle_bloc)
                stats = periode['stats'] if periode else None
                rmr = calculer_rmr(stats) if stats else None
                item = QTableWidgetItem(f"{rmr:.2f}" if rmr is not None else "—")
                item.setBackground(COULEURS_BLOCS[cle_bloc])
                self.table.setItem(row, col, item)
                col += 1

    # ── Suppression / remise à zéro ───────────────────────────────────────────

    def _supprimer_selection(self):
        lignes = sorted(
            {idx.row() for idx in self.table.selectedIndexes()},
            reverse=True,
        )
        if not lignes:
            return
        noms = list(self.resume_data.keys())
        for row in lignes:
            if row < len(noms):
                del self.resume_data[noms[row]]
        self._actualiser_tableau()

    def _vider_resume(self):
        if not self.resume_data:
            return
        rep = QMessageBox.question(
            self, "Vider le résumé",
            "Supprimer tous les examens du résumé ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if rep == QMessageBox.StandardButton.Yes:
            self.resume_data.clear()
            self._actualiser_tableau()

    # ── Export Excel ───────────────────────────────────────────────────────────

    def _exporter_excel(self):
        if not self.resume_data:
            QMessageBox.information(self, "Export", "Le résumé est vide. Envoyez d'abord des examens.")
            return

        chemin, _ = QFileDialog.getSaveFileName(
            self, "Exporter le résumé", "resume_lapec.xlsx",
            "Classeur Excel (*.xlsx)"
        )
        if not chemin:
            return

        try:
            self._ecrire_excel(chemin)
            QMessageBox.information(self, "Export réussi", f"Fichier enregistré :\n{chemin}")
        except Exception as exc:
            QMessageBox.critical(self, "Erreur d'export", str(exc))

    def _ecrire_excel(self, chemin: str):
        noms = list(self.resume_data.keys())
        vars_bloc = self._vars_bloc()
        col_headers = self._colonnes()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résumé"

        font_header = Font(bold=True)
        align_centre = Alignment(horizontal='center')

        # Fills par bloc (et colonnes sans bloc)
        fills = {cle: PatternFill('solid', fgColor=hex_) for cle, hex_ in COULEURS_EXCEL.items()}
        fill_vide = PatternFill(fill_type=None)

        # Correspondance index de colonne (0-based) → cle_bloc (ou None)
        col_bloc: list[str | None] = [None, None]  # Examen, Visite
        for cle_bloc, _ in BLOCS_RESUME:
            col_bloc += [cle_bloc] * len(vars_bloc)
        for cle_bloc, _ in BLOCS_RESUME:
            col_bloc.append(cle_bloc)  # RMR

        # En-têtes (ligne 1)
        for c_idx, (header, bloc) in enumerate(zip(col_headers, col_bloc), start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = font_header
            cell.alignment = align_centre
            if bloc:
                cell.fill = fills[bloc]

        # Données (lignes 2+)
        for row_idx, nom in enumerate(noms, start=2):
            res = self.resume_data[nom]
            col = 0

            # Examen
            ws.cell(row=row_idx, column=col + 1, value=nom)
            col += 1

            # Visite
            visite = deduire_visite(nom)
            ws.cell(row=row_idx, column=col + 1, value=visite)
            col += 1

            # Blocs de variables
            for cle_bloc, _ in BLOCS_RESUME:
                periode = res.get(cle_bloc)
                stats = periode['stats'] if periode else {}
                fill = fills[cle_bloc]
                for cle_stat, _, decimales in vars_bloc:
                    val = stats.get(cle_stat)
                    cell = ws.cell(row=row_idx, column=col + 1)
                    if val is not None:
                        cell.value = round(val, decimales)
                        cell.number_format = '0.000' if decimales == 3 else '0.00'
                    cell.fill = fill
                    col += 1

            # RMR par bloc
            for cle_bloc, _ in BLOCS_RESUME:
                periode = res.get(cle_bloc)
                stats = periode['stats'] if periode else None
                rmr = calculer_rmr(stats) if stats else None
                cell = ws.cell(row=row_idx, column=col + 1)
                if rmr is not None:
                    cell.value = rmr
                    cell.number_format = '0.00'
                cell.fill = fills[cle_bloc]
                col += 1

        # Largeur approximative des colonnes
        for c_idx, header in enumerate(col_headers, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max(10, len(header) + 2)

        # Figer la première colonne et la ligne d'en-tête
        ws.freeze_panes = 'C2'

        wb.save(chemin)

    # ── Styles ─────────────────────────────────────────────────────────────────

    def _style_btn_danger(self):
        return """
            QPushButton {
                background-color: #e05c3a; color: white; border: none;
                padding: 8px 18px; border-radius: 4px; font-size: 13px;
            }
            QPushButton:hover { background-color: #c44e30; }
            QPushButton:pressed { background-color: #a83e24; }
        """

    def _style_btn_vert(self):
        return """
            QPushButton {
                background-color: #1d6a3a; color: white; border: none;
                padding: 8px 18px; border-radius: 4px; font-size: 13px;
            }
            QPushButton:hover { background-color: #175730; }
            QPushButton:pressed { background-color: #124526; }
        """
